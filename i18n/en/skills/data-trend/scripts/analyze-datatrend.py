#!/usr/bin/env python3
"""
ABA个训数据趋势分析脚本
从每日反馈表PDF中提取个训项目数据，生成趋势分析Excel报告。

用法：
    python3 analyze.py --files file1.pdf file2.pdf ... --output report.xlsx

依赖：
    - poppler-utils (提供 pdftotext 命令)
    - openpyxl (生成Excel)
"""

import argparse
import re
import subprocess
import sys
import os
from collections import OrderedDict
from itertools import groupby

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：需要安装 openpyxl。请运行：pip install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════
# 1. PDF文本提取
# ═══════════════════════════════════════════════════════════════════════

def _find_pdftotext():
    """查找支持CJK的pdftotext路径（优先scoop版本）"""
    # Windows scoop安装的poppler版本支持CJK
    scoop_path = os.path.expanduser('~/scoop/shims/pdftotext.exe')
    if os.path.exists(scoop_path):
        return scoop_path
    scoop_path2 = os.path.expanduser('~/scoop/apps/poppler/current/Library/bin/pdftotext.exe')
    if os.path.exists(scoop_path2):
        return scoop_path2
    return 'pdftotext'

PDFTOTEXT_CMD = _find_pdftotext()


def check_pdftotext():
    """检查pdftotext是否可用"""
    try:
        subprocess.run([PDFTOTEXT_CMD, '-v'], capture_output=True, timeout=5)
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def extract_text(pdf_path):
    """用pdftotext提取PDF文本（不加-layout参数，逐行输出更适合表格解析）"""
    if not os.path.exists(pdf_path):
        print(f"  警告：文件不存在 {pdf_path}")
        return ""
    result = subprocess.run(
        [PDFTOTEXT_CMD, pdf_path, '-'],
        capture_output=True, text=True, encoding='utf-8', errors='replace', timeout=30
    )
    if result.returncode != 0:
        print(f"  警告：pdftotext 提取失败 {pdf_path}: {result.stderr}")
        return ""
    return result.stdout


def extract_date_from_filename(filepath):
    """从文件名中提取日期，格式如 LQL20260316"""
    basename = os.path.basename(filepath)
    # 匹配8位数字日期
    m = re.search(r'(\d{4})(\d{2})(\d{2})', basename)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None


def extract_date_from_text(text):
    """从PDF文本内容中提取日期"""
    m = re.search(r'日期[：:]\s*(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', text)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None


def extract_student_name(text):
    """从PDF文本中提取学生姓名"""
    if not text:
        return None
    m = re.search(r'学生姓名[：:]\s*([\u4e00-\u9fff]+)', text)
    if m:
        return m.group(1)
    return None


# ═══════════════════════════════════════════════════════════════════════
# 2. 个训项目数据解析（状态机解析器）
# ═══════════════════════════════════════════════════════════════════════

def parse_individual_training_items(text):
    """
    从反馈表文本中解析个训项目数据。
    
    使用状态机逐行解析，因为pdftotext输出的格式是每个字段单独一行：
        项目名称（可能跨多行）
        百分比%
        是/否
        下一个项目名称...
    
    返回：(items_dict, mastery_standard)
        items_dict: {项目名称: {'pct': int或None, 'mastered': '是'/'否'/None}}
        mastery_standard: "80%" 或 "90%"
    """
    items = {}
    lines = text.split('\n')
    
    # 定位个训项目区域的起止位置
    start_idx = None
    end_idx = None
    mastery_standard = "80%"
    
    for i, line in enumerate(lines):
        if '掌握标准' in line:
            start_idx = i + 1
            if '90%' in line:
                mastery_standard = "90%"
            elif '80%' in line:
                mastery_standard = "80%"
        # 个训项目区域在这些标记之前结束
        if start_idx is not None and any(marker in line for marker in [
            '负面行为', '流利度项目', '融合干预'
        ]):
            # 流利度项目也是我们需要单独处理的区域，先在这里截断多回合项目
            end_idx = i
            break
    
    if start_idx is None:
        return items, mastery_standard
    if end_idx is None:
        end_idx = len(lines)
    
    # 提取区域内的有效行（去掉表头和空行）
    section_lines = lines[start_idx:end_idx]
    tokens = []
    for line in section_lines:
        s = line.strip()
        if s and s not in ['项目名称', '%', '是否掌握'] \
                and '项目名称' not in s and '是否掌握' not in s \
                and s != '%':
            tokens.append(s)
    
    # 状态机解析
    # token序列的模式是：[名称部分...] [百分比%或无] [是/否/无]
    current_name_parts = []
    i = 0
    
    while i < len(tokens):
        token = tokens[i]
        
        # 情况1：token是百分比值（如 "17%"、"100%"、"0%"）
        pct_match = re.match(r'^(\d+)%$', token)
        if pct_match:
            name = ''.join(current_name_parts)
            pct = int(pct_match.group(1))
            mastered = None
            # 下一个token应该是 是/否
            if i + 1 < len(tokens) and tokens[i + 1] in ['是', '否']:
                mastered = tokens[i + 1]
                i += 2
            else:
                i += 1
            if name:
                items[name] = {'pct': pct, 'mastered': mastered}
            current_name_parts = []
            continue
        
        # 情况2：token是"无"（项目当天未做）
        if token == '无':
            if i + 1 < len(tokens) and tokens[i + 1] == '无':
                # "无 无" = 百分比无 + 是否掌握无
                name = ''.join(current_name_parts)
                if name:
                    items[name] = {'pct': None, 'mastered': None}
                current_name_parts = []
                i += 2
                continue
            elif current_name_parts:
                # 单个"无"作为值
                name = ''.join(current_name_parts)
                if name:
                    items[name] = {'pct': None, 'mastered': None}
                current_name_parts = []
                i += 1
                continue
        
        # 情况3：token是孤立的 是/否（前面没有名称积累，可能是解析残留）
        if token in ['是', '否'] and not current_name_parts:
            i += 1
            continue
        
        # 情况4：其他文本 -> 累积为项目名称的一部分
        current_name_parts.append(token)
        i += 1
    
    return items, mastery_standard


def parse_fluency_items(text):
    """
    解析流利度项目数据。
    流利度项目格式不同，包含：项目名称、当日阶段、当日目标、最好数据、是否达标
    
    返回：list of dicts
    """
    fluency_items = []
    lines = text.split('\n')
    
    in_fluency = False
    for i, line in enumerate(lines):
        if '流利度项目' in line:
            in_fluency = True
            continue
        if in_fluency and any(marker in line for marker in ['融合干预', '负面行为']):
            break
    
    # 流利度项目的解析比较简单，通常只有1-2个项目
    # 由于格式复杂且项目少，这里做简化处理
    # TODO: 完整的流利度项目解析
    
    return fluency_items


# ═══════════════════════════════════════════════════════════════════════
# 3. 项目名称标准化（跨天匹配）
# ═══════════════════════════════════════════════════════════════════════

def canonicalize_name(name):
    """
    标准化项目名称，用于跨天匹配同一项目。
    
    处理常见变体：
    - 书名号：社交故事《分享》 → 社交故事-分享
    - 全角/半角括号统一
    - 去除多余空格
    - 统一连接符
    """
    n = name
    n = n.replace('《', '-').replace('》', '')
    n = n.replace('（', '(').replace('）', ')')
    n = re.sub(r'\s+', '', n)
    n = re.sub(r'[-—]+', '-', n)
    # 去除末尾的连接符
    n = n.strip('-')
    return n


def categorize_item(name):
    """根据项目名称自动归类到领域"""
    n = name
    if any(k in n for k in ['语义网络', '仿说', '说出类别', '命名']):
        return '语言/认知'
    if any(k in n for k in ['听者', '两步指令', '多步骤指令']):
        return '听理解'
    if any(k in n for k in ['社交故事', '社交游戏', '眼神', '炫耀', '社交技能']):
        return '社交'
    if any(k in n for k in ['记忆', '对话', '每日回忆']):
        return '认知/对话'
    if any(k in n for k in ['切水果', '捏夹子', '控笔', '流利度']):
        return '精细/流利度'
    return '其他'


# ═══════════════════════════════════════════════════════════════════════
# 4. 趋势分析
# ═══════════════════════════════════════════════════════════════════════

def calculate_trend(values):
    """
    根据多日数据计算趋势方向。
    
    values: list of int或None（None表示当天未做）
    返回: 趋势描述字符串
    """
    valid = [v for v in values if v is not None]
    
    if len(valid) < 2:
        return "数据不足"
    
    # 分前后两半比较
    mid = len(valid) // 2
    first_half = valid[:mid]
    second_half = valid[mid:]
    avg_first = sum(first_half) / len(first_half)
    avg_second = sum(second_half) / len(second_half)
    diff = avg_second - avg_first
    
    if all(v >= 80 for v in valid):
        return "★ 稳定高位"
    elif all(v <= 20 for v in valid):
        return "▽ 持续低位"
    elif diff > 15:
        return "↑ 上升"
    elif diff < -15:
        return "↓ 下降"
    else:
        return "～ 波动"


# ═══════════════════════════════════════════════════════════════════════
# 5. Excel报告生成
# ═══════════════════════════════════════════════════════════════════════

def generate_excel(results, dates, student_name, output_path):
    """生成Excel趋势分析报告"""
    wb = Workbook()
    
    # ─── Sheet 1: 数据表格 ───
    ws = wb.active
    ws.title = "个训数据趋势"
    
    # 样式定义
    header_font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cat_fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    cat_font = Font(name='Arial', bold=True, size=11)
    data_font = Font(name='Arial', size=10)
    mastered_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    low_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    near_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    rising_font = Font(name='Arial', size=10, color='006100')
    falling_font = Font(name='Arial', size=10, color='9C0006')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # 标题
    num_cols = 2 + len(dates) + 3  # 领域 + 项目 + 日期列 + 均值 + 趋势 + 掌握状态
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws['A1'] = f'个训数据趋势分析 - {student_name}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    ws['A2'] = f'分析周期：{dates[0]} 至 {dates[-1]}（共{len(dates)}天数据）'
    ws['A2'].font = Font(name='Arial', size=10, color='595959')
    
    # 表头
    row = 4
    headers = ['领域', '项目名称'] + [d[-5:] for d in dates] + ['均值', '趋势', '掌握状态']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # 列宽
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 42
    for i in range(3, 3 + len(dates)):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions[get_column_letter(3 + len(dates))].width = 8
    ws.column_dimensions[get_column_letter(4 + len(dates))].width = 14
    ws.column_dimensions[get_column_letter(5 + len(dates))].width = 12
    
    # 数据行（按领域分组）
    results_sorted = sorted(results, key=lambda x: (x['category'], x['name']))
    row = 5
    current_cat = None
    
    for r in results_sorted:
        if r['category'] != current_cat:
            current_cat = r['category']
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
            cell = ws.cell(row=row, column=1, value=f"【{current_cat}】")
            cell.font = cat_font
            cell.fill = cat_fill
            for c in range(1, num_cols + 1):
                ws.cell(row=row, column=c).fill = cat_fill
                ws.cell(row=row, column=c).border = thin_border
            row += 1
        
        # 领域
        ws.cell(row=row, column=1, value=r['category']).font = data_font
        ws.cell(row=row, column=1).border = thin_border
        
        # 项目名称
        ws.cell(row=row, column=2, value=r['name']).font = data_font
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=2).border = thin_border
        
        # 各日数据
        for i, v in enumerate(r['values']):
            cell = ws.cell(row=row, column=3 + i)
            if v is not None:
                cell.value = v / 100
                cell.number_format = '0%'
                if v >= 80:
                    cell.fill = mastered_fill
                elif v <= 20:
                    cell.fill = low_fill
            else:
                cell.value = '--'
                cell.alignment = Alignment(horizontal='center')
            cell.font = data_font
            cell.border = thin_border
        
        # 均值
        avg_cell = ws.cell(row=row, column=3 + len(dates))
        if r['avg'] is not None:
            avg_cell.value = r['avg'] / 100
            avg_cell.number_format = '0%'
        else:
            avg_cell.value = '--'
        avg_cell.font = data_font
        avg_cell.border = thin_border
        
        # 趋势
        trend_cell = ws.cell(row=row, column=4 + len(dates), value=r['trend'])
        if '上升' in r['trend']:
            trend_cell.font = rising_font
        elif '下降' in r['trend']:
            trend_cell.font = falling_font
        else:
            trend_cell.font = data_font
        trend_cell.border = thin_border
        
        # 掌握状态
        ms_cell = ws.cell(row=row, column=5 + len(dates))
        if r['mastered_days']:
            ms_cell.value = f"已掌握 ({r['mastered_days'][0][-5:]})"
            ms_cell.fill = mastered_fill
        elif r['avg'] is not None and r['avg'] >= 70:
            ms_cell.value = "接近掌握"
            ms_cell.fill = near_fill
        elif r['avg'] is not None and r['avg'] <= 20:
            ms_cell.value = "需关注"
            ms_cell.fill = low_fill
        else:
            ms_cell.value = "进行中"
        ms_cell.font = data_font
        ms_cell.border = thin_border
        
        row += 1
    
    # ─── Sheet 2: 文字总结 ───
    ws2 = wb.create_sheet("文字总结")
    ws2.column_dimensions['A'].width = 90
    
    summary = generate_text_summary(results, dates, student_name)
    for i, line in enumerate(summary.split('\n'), 1):
        cell = ws2.cell(row=i, column=1, value=line)
        if any(line.startswith(p) for p in ['【', '一、', '二、', '三、', '四、', '五、', '六、']):
            cell.font = Font(name='Arial', bold=True, size=11)
        else:
            cell.font = Font(name='Arial', size=10)
    
    # 保存
    wb.save(output_path)
    return summary


# ═══════════════════════════════════════════════════════════════════════
# 6. 文字总结生成
# ═══════════════════════════════════════════════════════════════════════

def generate_text_summary(results, dates, student_name):
    """生成可直接贴进报告的中文文字总结"""
    lines = []
    lines.append(f"【个训数据趋势分析】")
    lines.append(f"分析周期：{dates[0]} 至 {dates[-1]}（共{len(dates)}天数据）")
    lines.append(f"学生：{student_name}")
    lines.append("")
    
    # 分类统计
    mastered = [r for r in results if r['mastered_days']]
    near = [r for r in results if r['avg'] is not None and r['avg'] >= 70 and not r['mastered_days']]
    rising = [r for r in results if '上升' in r['trend']]
    falling = [r for r in results if '下降' in r['trend']]
    low = [r for r in results if r['avg'] is not None and r['avg'] <= 20]
    
    # 一、掌握项目
    lines.append("一、本周期掌握项目")
    if mastered:
        for m in mastered:
            vals_str = ' → '.join(f"{v}%" for v in m['values'] if v is not None)
            lines.append(f"  ✓ {m['name']}（掌握日期：{m['mastered_days'][0]}，数据走向：{vals_str}）")
    else:
        lines.append("  （本周期无新掌握项目）")
    
    # 二、接近掌握
    lines.append("")
    lines.append("二、接近掌握的项目（均值≥70%，需继续保持）")
    if near:
        for m in near:
            lines.append(f"  → {m['name']}（均值{m['avg']:.0f}%）")
    else:
        lines.append("  （无）")
    
    # 三、上升趋势
    lines.append("")
    lines.append("三、进步明显的项目（上升趋势）")
    if rising:
        for m in rising:
            vals_str = ' → '.join(f"{v}%" for v in m['values'] if v is not None)
            lines.append(f"  ↑ {m['name']}（{vals_str}）")
    else:
        lines.append("  （无）")
    
    # 四、下降趋势
    lines.append("")
    lines.append("四、下降趋势项目（需检视教学策略）")
    if falling:
        for m in falling:
            vals_str = ' → '.join(f"{v}%" for v in m['values'] if v is not None)
            lines.append(f"  ↓ {m['name']}（{vals_str}）")
    else:
        lines.append("  （无）")
    
    # 五、持续低位
    lines.append("")
    lines.append("五、需要关注的项目（持续低位，均值≤20%）")
    if low:
        for m in low:
            lines.append(f"  ⚠ {m['name']}（均值{m['avg']:.0f}%）")
    else:
        lines.append("  （无）")
    
    # 六、数据概览
    lines.append("")
    lines.append("六、数据概览")
    lines.append(f"  追踪项目总数：{len(results)}个")
    lines.append(f"  已掌握：{len(mastered)}个")
    lines.append(f"  接近掌握：{len(near)}个")
    lines.append(f"  上升趋势：{len(rising)}个")
    lines.append(f"  下降趋势：{len(falling)}个")
    lines.append(f"  持续低位需关注：{len(low)}个")
    
    return '\n'.join(lines)


# ═══════════════════════════════════════════════════════════════════════
# 7. 主流程
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='ABA个训数据趋势分析 - 从每日反馈表PDF生成趋势报告'
    )
    parser.add_argument(
        '--files', nargs='+', required=True,
        help='每日反馈表PDF文件路径（可多个）'
    )
    parser.add_argument(
        '--output', default='个训数据趋势分析.xlsx',
        help='输出Excel文件路径（默认：个训数据趋势分析.xlsx）'
    )
    parser.add_argument(
        '--student', default=None,
        help='学生姓名（可选，默认从PDF中自动提取）'
    )
    args = parser.parse_args()
    
    # 检查依赖
    if not check_pdftotext():
        print("错误：未找到 pdftotext 工具。请安装：sudo apt-get install -y poppler-utils")
        sys.exit(1)
    
    print(f"📋 ABA个训数据趋势分析")
    print(f"   输入文件：{len(args.files)} 个PDF")
    print(f"   输出路径：{args.output}")
    print()
    
    # ─── 提取数据 ───
    all_data = OrderedDict()
    canon_map = {}  # canonical_name -> first_seen_display_name
    student_name = args.student
    
    # 先提取日期并排序
    file_dates = []
    for fp in args.files:
        date = extract_date_from_filename(fp)
        if date is None:
            text = extract_text(fp)
            date = extract_date_from_text(text)
        if date is None:
            print(f"  ⚠ 无法提取日期：{fp}，跳过")
            continue
        file_dates.append((fp, date))
    
    # 按日期排序
    file_dates.sort(key=lambda x: x[1])
    
    for fp, date in file_dates:
        print(f"  📄 处理 {os.path.basename(fp)} ({date})...")
        text = extract_text(fp)
        
        # 提取学生姓名（如果还没有）
        if student_name is None:
            student_name = extract_student_name(text)
        
        # 解析个训项目
        items, mastery_std = parse_individual_training_items(text)
        print(f"     提取到 {len(items)} 个项目（掌握标准：{mastery_std}）")
        
        # 标准化名称并存储
        day_data = {}
        for name, data in items.items():
            cn = canonicalize_name(name)
            if cn not in canon_map:
                canon_map[cn] = name
            day_data[cn] = data
        
        all_data[date] = day_data
    
    if not all_data:
        print("错误：没有成功提取到任何数据")
        sys.exit(1)
    
    if student_name is None:
        student_name = "未知学生"
    
    dates = list(all_data.keys())
    all_canons = sorted(set().union(*(d.keys() for d in all_data.values())))
    
    print(f"\n📊 分析 {len(all_canons)} 个项目，跨 {len(dates)} 天")
    
    # ─── 趋势分析 ───
    results = []
    for cn in all_canons:
        display_name = canon_map[cn]
        category = categorize_item(display_name)
        
        values = []
        mastered_days = []
        for date in dates:
            item_data = all_data[date].get(cn, {})
            values.append(item_data.get('pct'))
            if item_data.get('mastered') == '是':
                mastered_days.append(date)
        
        valid = [v for v in values if v is not None]
        avg = sum(valid) / len(valid) if valid else None
        trend = calculate_trend(values)
        
        results.append({
            'name': display_name,
            'canon': cn,
            'category': category,
            'values': values,
            'avg': avg,
            'trend': trend,
            'mastered_days': mastered_days,
        })
    
    # ─── 生成报告 ───
    print(f"\n📝 生成Excel报告...")
    summary = generate_excel(results, dates, student_name, args.output)
    
    print(f"\n✅ 报告已保存至：{args.output}")
    print(f"\n{'='*60}")
    print(summary)
    print(f"{'='*60}")


if __name__ == '__main__':
    main()
