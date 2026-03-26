#!/usr/bin/env python3
"""
ABA融合数据对比与IEP达标判定脚本
从融合每日反馈表PDF中提取融合干预数据，进行趋势分析，
可选与IEP目标进行达标判定，并在月度模式下输出入园评估全局扫描。

用法：
    # 日常模式
    python3 fusion_analyze.py --files f1.pdf f2.pdf ... --output report.xlsx

    # 月度模式 + IEP对比
    python3 fusion_analyze.py --files f1.pdf f2.pdf ... --iep iep.docx --mode monthly --output report.xlsx

依赖：
    - poppler-utils (pdftotext)
    - openpyxl
    - pandoc (IEP docx解析，可选)
"""

import argparse
import re
import subprocess
import sys
import os
from collections import OrderedDict

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：需要安装 openpyxl。请运行：pip install openpyxl")
    sys.exit(1)


# ═══════════════════════════════════════════════════════════════════════
# 1. 工具函数
# ═══════════════════════════════════════════════════════════════════════

def _find_pdftotext():
    """查找支持CJK的pdftotext路径（优先scoop版本）"""
    scoop_path = os.path.expanduser('~/scoop/shims/pdftotext.exe')
    if os.path.exists(scoop_path):
        return scoop_path
    scoop_path2 = os.path.expanduser('~/scoop/apps/poppler/current/Library/bin/pdftotext.exe')
    if os.path.exists(scoop_path2):
        return scoop_path2
    return 'pdftotext'

PDFTOTEXT_CMD = _find_pdftotext()

def check_pdftotext():
    try:
        subprocess.run([PDFTOTEXT_CMD, '-v'], capture_output=True, timeout=5)
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False

def extract_layout(pdf_path):
    """用layout模式提取PDF文本，保留表格列对齐"""
    result = subprocess.run(
        [PDFTOTEXT_CMD, '-layout', pdf_path, '-'],
        capture_output=True, text=True, encoding='utf-8', errors='replace', timeout=30
    )
    return result.stdout

def extract_plain(pdf_path):
    """用纯文本模式提取PDF（个训项目解析用）"""
    result = subprocess.run(
        [PDFTOTEXT_CMD, pdf_path, '-'],
        capture_output=True, text=True, encoding='utf-8', errors='replace', timeout=30
    )
    return result.stdout

def extract_date_from_filename(filepath):
    basename = os.path.basename(filepath)
    m = re.search(r'(\d{4})(\d{2})(\d{2})', basename)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    return None

def extract_date_from_text(text):
    m = re.search(r'日期[：:]\s*(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日', text)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None

def extract_student_name(text):
    if not text:
        return None
    m = re.search(r'学生姓名[：:]\s*([\u4e00-\u9fff]+)', text)
    return m.group(1) if m else None

def calculate_trend(values):
    valid = [v for v in values if v is not None]
    if len(valid) < 2:
        return "数据不足"
    mid = len(valid) // 2
    avg1 = sum(valid[:mid]) / len(valid[:mid])
    avg2 = sum(valid[mid:]) / len(valid[mid:])
    diff = avg2 - avg1
    if all(v >= 80 for v in valid):
        return "★ 稳定高位"
    elif all(v <= 20 for v in valid):
        return "▽ 持续低位"
    elif diff > 15:
        return "↑ 上升"
    elif diff < -15:
        return "↓ 下降"
    else:
        return "～ 波动"


# ═══════════════════════════════════════════════════════════════════════
# 2. 个训项目解析（layout模式，支持双列表格）
# ═══════════════════════════════════════════════════════════════════════

def _parse_layout_column(raw_lines):
    """解析单列layout文本，处理跨行项目名称。

    layout模式每行格式为：
      项目名称(可能不完整)     百分比%     是/否
      项目名称续行（下一行，无百分比）
    """
    items = {}
    current_name_parts = []
    pending_pct = None
    pending_mastered = None
    has_pending = False

    for line in raw_lines:
        stripped = line.strip()
        if not stripped or '项目名称' in stripped or '是否掌握' in stripped or stripped == '%':
            continue

        # 匹配：名称  数字%  是/否
        m = re.match(r'^(.+?)\s{2,}(\d+)%\s{2,}(是|否)\s*$', stripped)
        if m:
            if has_pending and current_name_parts:
                name = ''.join(current_name_parts)
                items[name] = {'pct': pending_pct, 'mastered': pending_mastered}
            current_name_parts = [m.group(1).strip()]
            pending_pct = int(m.group(2))
            pending_mastered = m.group(3)
            has_pending = True
            continue

        # 匹配：名称  无  无
        m2 = re.match(r'^(.+?)\s{2,}无\s{2,}无\s*$', stripped)
        if m2:
            if has_pending and current_name_parts:
                name = ''.join(current_name_parts)
                items[name] = {'pct': pending_pct, 'mastered': pending_mastered}
            current_name_parts = [m2.group(1).strip()]
            pending_pct = None
            pending_mastered = None
            has_pending = True
            continue

        # 续行：追加到当前项目名称
        current_name_parts.append(stripped)

    # 刷出最后一个项目
    if has_pending and current_name_parts:
        name = ''.join(current_name_parts)
        items[name] = {'pct': pending_pct, 'mastered': pending_mastered}

    return items


def parse_individual_training(layout_text):
    """从layout模式文本提取个训项目数据（支持双列表格）"""
    items = {}
    lines = layout_text.split('\n')
    start_idx = None
    end_idx = None
    mastery_std = "80%"

    for i, line in enumerate(lines):
        if '掌握标准' in line and start_idx is None:
            start_idx = i + 1
            if '90%' in line: mastery_std = "90%"
        elif start_idx and any(m in line for m in ['负面行为', '流利度项目', '融合干预']):
            end_idx = i
            break

    if start_idx is None:
        return items, mastery_std
    if end_idx is None:
        end_idx = len(lines)

    section = lines[start_idx:end_idx]

    # 找分列点：第二个"项目名称"的起始位置
    split_col = None
    for line in section:
        positions = [m.start() for m in re.finditer(r'项目名称', line)]
        if len(positions) >= 2:
            split_col = positions[1]
            break

    if split_col is None:
        # 单列模式：直接解析全部行
        return _parse_layout_column(section), mastery_std

    # 双列分离
    left_lines = []
    right_lines = []
    for line in section:
        left_part = line[:split_col]
        right_part = line[split_col:] if len(line) > split_col else ''
        if left_part.strip():
            left_lines.append(left_part)
        if right_part.strip():
            right_lines.append(right_part)

    left_items = _parse_layout_column(left_lines)
    right_items = _parse_layout_column(right_lines)
    items.update(left_items)
    items.update(right_items)
    return items, mastery_std


# ═══════════════════════════════════════════════════════════════════════
# 3. 融合干预数据解析（layout模式）
# ═══════════════════════════════════════════════════════════════════════

def _find_nums_and_pcts(line):
    """从一行中提取所有 X次 和 X% 和 无"""
    parts = re.findall(r'(\d+)次|(\d+)%|(无)', line)
    result = []
    for p in parts:
        if p[0]:
            result.append(('n', int(p[0])))
        elif p[1]:
            result.append(('p', int(p[1])))
        elif p[2]:
            result.append(('x', None))
    return result

def _extract_triplet(nums, offset=0):
    """从nums列表的offset位置提取(独立, 辅助, 百分比)三元组"""
    if offset + 2 >= len(nums):
        return None, None, None
    ind = nums[offset][1] if nums[offset][0] == 'n' else None
    ast = nums[offset + 1][1] if nums[offset + 1][0] == 'n' else None
    pct = nums[offset + 2][1] if nums[offset + 2][0] == 'p' else None
    return ind, ast, pct


def parse_peer_intervention(layout_text):
    """解析同伴介入数据"""
    peer = {}
    lines = layout_text.split('\n')

    for line in lines:
        if '游戏发起' in line:
            nums = _find_nums_and_pcts(line)
            if len(nums) >= 3:
                ind, ast, pct = _extract_triplet(nums, 0)
                peer['游戏发起'] = {'独立': ind, '辅助': ast, 'pct': pct}
            if len(nums) >= 6:
                ind, ast, pct = _extract_triplet(nums, 3)
                peer['游戏回应'] = {'独立': ind, '辅助': ast, 'pct': pct}

        if '影子游戏' in line:
            nums = _find_nums_and_pcts(line)
            if len(nums) >= 3:
                ind, ast, pct = _extract_triplet(nums, 0)
                peer['影子游戏'] = {'独立': ind, '辅助': ast, 'pct': pct}
            if len(nums) >= 6:
                ind, ast, pct = _extract_triplet(nums, 3)
                peer['抓捕小鱼'] = {'独立': ind, '辅助': ast, 'pct': pct}

    return peer


def parse_inclusion_data(layout_text):
    """解析随班融合数据（集教课关注率、指令回应等）"""
    inclusion = {}
    lines = layout_text.split('\n')

    for i, line in enumerate(lines):
        # 集教课关注率
        if '集教' in line and ('关注率' in line or '干预时' in line):
            ctx = ' '.join(lines[max(0, i - 1):min(len(lines), i + 3)])
            times = re.findall(r'(\d+)分(\d+)秒', ctx)
            pcts = re.findall(r'(\d+)%', ctx)
            if times and pcts:
                total = int(times[0][0]) * 60 + int(times[0][1])
                attn = int(times[1][0]) * 60 + int(times[1][1]) if len(times) >= 2 else None
                inclusion['集教课关注率'] = {
                    'total_secs': total,
                    'attention_secs': attn,
                    'pct': int(pcts[0])
                }

        # 拆分单独指令 (may span two lines: "集教课-拆分" then "单独指令")
        if '拆分' in line:
            ctx = ' '.join(lines[max(0, i - 1):min(len(lines), i + 3)])
            nums = re.findall(r'(\d+)次', ctx)
            pcts = re.findall(r'(\d+)%', ctx)
            if len(nums) >= 2 and '拆分单独指令' not in inclusion:
                t, r_ = int(nums[0]), int(nums[1])
                inclusion['拆分单独指令'] = {
                    'total': t, 'response': r_,
                    'pct': int(pcts[0]) if pcts else (round(r_ / t * 100) if t > 0 else 0)
                }

        # 转衔-户外跟随走 (may span: "转衔-户外" then "跟随走")
        if '转衔' in line:
            ctx = ' '.join(lines[max(0, i):min(len(lines), i + 2)])
            nums = re.findall(r'(\d+)次', ctx)
            pcts = re.findall(r'(\d+)%', ctx)
            if len(nums) >= 2:
                t, r_ = int(nums[0]), int(nums[1])
                inclusion['转衔跟随'] = {
                    'total': t, 'response': r_,
                    'pct': int(pcts[0]) if pcts else (round(r_ / t * 100) if t > 0 else 0)
                }
            elif '无' in ctx and not nums:
                inclusion['转衔跟随'] = {'total': None, 'response': None, 'pct': None}

        # 回应班级集体指令
        if '回应班级集' in line:
            ctx = ' '.join(lines[max(0, i):min(len(lines), i + 4)])
            nums = re.findall(r'(\d+)次', ctx)
            pcts = re.findall(r'(\d+)%', ctx)
            if len(nums) >= 2 and pcts:
                # 用百分比来判断哪个是total哪个是response
                pct_val = int(pcts[0])
                n1, n2 = int(nums[0]), int(nums[1])
                # 如果 n2/n1 接近 pct_val/100，则 n1=total, n2=response
                # 否则反过来
                if n1 > 0 and abs(n2 / n1 * 100 - pct_val) < 15:
                    inclusion['集体指令'] = {'total': n1, 'response': n2, 'pct': pct_val}
                elif n2 > 0 and abs(n1 / n2 * 100 - pct_val) < 15:
                    inclusion['集体指令'] = {'total': n2, 'response': n1, 'pct': pct_val}
                else:
                    inclusion['集体指令'] = {'total': n1, 'response': n2, 'pct': pct_val}
            elif len(nums) >= 2:
                t, r_ = int(nums[0]), int(nums[1])
                inclusion['集体指令'] = {
                    'total': t, 'response': r_,
                    'pct': round(r_ / t * 100) if t > 0 else 0
                }

        # 向班级老师或同伴提要求
        if '向班级老师' in line:
            ctx = ' '.join(lines[max(0, i):min(len(lines), i + 3)])
            nums = re.findall(r'(\d+)次', ctx)
            pcts = re.findall(r'(\d+)%', ctx)
            if len(nums) >= 2:
                t, r_ = int(nums[0]), int(nums[1])
                inclusion['向老师提要求'] = {
                    'total': t, 'response': r_,
                    'pct': int(pcts[0]) if pcts else (round(r_ / t * 100) if t > 0 else 0)
                }

    return inclusion


def parse_usopac(layout_text, student_name=None):
    """解析USOPAC社交观察数据。student_name用于识别个案行。"""
    usopac = {'个案': {}, '同伴': {}}
    lines = layout_text.split('\n')
    categories = ['U独自空闲', 'S独自游戏', 'O观察者', 'P平行游戏',
                  'A联合游戏', 'C合作游戏', '独立发起', '独立回应']

    # 找到USOPAC区域
    usopac_start = None
    for i, line in enumerate(lines):
        if 'USOPAC' in line:
            usopac_start = i
            break

    if usopac_start is None:
        return usopac

    # 在USOPAC区域后找个案行和同伴行
    # 个案行：包含学生姓名或昵称的行（如果不知道姓名，取USOPAC后第一个非表头数据行）
    # 同伴行：紧跟个案行之后包含"同伴"的行
    found_student = False
    for i in range(usopac_start + 1, min(usopac_start + 15, len(lines))):
        line = lines[i]

        # 跳过表头行
        if any(h in line for h in ['行为者', 'U独自', 'S独自', 'O观察', 'P平行']):
            continue

        is_student_line = False
        if student_name and student_name in line:
            is_student_line = True
        elif not found_student and '同伴' not in line and line.strip():
            # 如果不知道学生姓名，USOPAC后第一个非表头非同伴的数据行就是个案行
            nums = re.findall(r'(\d+)次|(无)', line)
            if nums:
                is_student_line = True

        if is_student_line and not found_student:
            found_student = True
            nums = re.findall(r'(\d+)次|(无)', line)
            vals = []
            for n in nums:
                vals.append(int(n[0]) if n[0] else None)
            for j, cat in enumerate(categories[:len(vals)]):
                usopac['个案'][cat] = vals[j]

        if '同伴' in line and found_student:
            nums = re.findall(r'(\d+)次|(无)', line)
            vals = []
            for n in nums:
                vals.append(int(n[0]) if n[0] else None)
            for j, cat in enumerate(categories[:len(vals)]):
                usopac['同伴'][cat] = vals[j]
            break

    return usopac


def parse_daily_notes(layout_text):
    """提取每日亮点和每日小结文字"""
    lines = layout_text.split('\n')
    highlight = ""
    summary = ""

    for i, line in enumerate(lines):
        if '每日亮点' in line:
            for j in range(i + 1, min(i + 5, len(lines))):
                l = lines[j].strip()
                if l and '每日小结' not in l and '版权所有' not in l:
                    highlight += l + " "
                if '每日小结' in l:
                    break
        if '每日小结' in line:
            for j in range(i + 1, min(i + 15, len(lines))):
                l = lines[j].strip()
                if l and '家长注意' not in l and '版权所有' not in l:
                    summary += l + " "
                if '家长注意' in l or '版权所有' in l:
                    break

    return highlight.strip(), summary.strip()


# ═══════════════════════════════════════════════════════════════════════
# 4. IEP目标解析
# ═══════════════════════════════════════════════════════════════════════

def parse_iep_goals(iep_path):
    """
    从IEP文档中提取目标。
    使用pandoc转markdown后，用正则+语义匹配提取。
    返回 list of dicts: [{'domain': '社交', 'goal': '...', 'target_pct': 80, 'metric': '...'}, ...]
    """
    goals = []

    try:
        result = subprocess.run(
            ['pandoc', iep_path, '-t', 'markdown'],
            capture_output=True, text=True, timeout=30
        )
        text = result.stdout
    except (FileNotFoundError, subprocess.TimeoutExpired):
        print("  警告：pandoc不可用或超时，跳过IEP解析")
        return goals

    # 提取目标：扫描全文找带百分比的目标描述
    # IEP格式差异大，用宽松匹配
    lines = text.split('\n')
    in_goals = False
    current_domain = None

    for i, line in enumerate(lines):
        # 检测领域
        stripped = re.sub(r'[\*\|#\[\]{}]', '', line).strip()
        if '长期目标' in line or '短期目标' in line or '学期目标' in line:
            in_goals = True

        if '集体' in stripped[:6] or ('集体' in stripped and '领域' in stripped):
            current_domain = '集体'
        elif '社交' in stripped[:6] or ('社交' in stripped and '领域' in stripped):
            current_domain = '社交'
        elif '自理' in stripped[:6] or ('自理' in stripped and ('领域' in stripped or '能力' in stripped)):
            current_domain = '自理'

        if not in_goals:
            continue

        # 找带百分比的目标行
        pct_matches = re.findall(r'(\d+)%', line)
        if not pct_matches or not current_domain:
            continue

        # 需要有动词性关键词
        has_verb = any(k in line for k in [
            '达到', '达', '稳定', '以上', '正确率', '关注率', '回应率',
            '跟随', '能够', '独立', '百分比'
        ])
        if not has_verb:
            continue

        # 清理文本
        goal_text = re.sub(r'[\*\|#\[\]{}]', '', line).strip()
        goal_text = re.sub(r'\s+', ' ', goal_text).strip()

        # 过滤太短或是表头/测量描述
        if len(goal_text) < 10:
            continue
        if any(k in goal_text for k in ['记录', '计算出百分比', '每月家长', '进展记录']):
            # 这行是测量方法描述，跳过——但如果同一行也包含目标，尝试提取目标部分
            # 典型格式："目标描述 测量方法 频率"
            # 只取第一个百分比前面的部分作为目标
            parts = re.split(r'记录|计算', goal_text)
            if parts[0].strip() and len(parts[0].strip()) > 10 and pct_matches:
                goal_text = parts[0].strip()
            else:
                continue

        target_pct = int(pct_matches[0])  # 第一个百分比通常是目标值
        goals.append({
            'domain': current_domain,
            'goal': goal_text,
            'target_pct': target_pct,
        })

    # 去重 + 过滤碎片
    seen = set()
    unique_goals = []
    for g in goals:
        # 过滤pandoc拆碎的残片（目标文本应该以中文字符开头，不以百分比或标点开头）
        gt = g['goal'].strip()
        if not gt:
            continue
        first_char = gt[0]
        if first_char in '达稳确分' and len(gt) < 30:
            continue  # 这些是被截断的行，如 "达到50%以上 数和辅助"
        if '<!--' in gt or '```' in gt:
            continue  # markdown残留

        key = gt[:25]
        if key not in seen:
            seen.add(key)
            unique_goals.append(g)

    return unique_goals


def match_goal_to_data(goal, fusion_data):
    """
    将IEP目标与融合数据进行匹配。
    返回 (current_value, data_source) 或 (None, None)
    """
    gt = goal['goal'].lower()

    # 集教课关注率
    if '关注率' in gt:
        d = fusion_data.get('随班融合', {}).get('集教课关注率')
        if d and d.get('pct') is not None:
            return d['pct'], '集教课关注率'

    # 集体指令回应率
    if '集体指令' in gt or ('指令' in gt and '回应' in gt and '集体' in gt):
        d = fusion_data.get('随班融合', {}).get('集体指令')
        if d and d.get('pct') is not None:
            return d['pct'], '集体指令回应率'

    # 拆分单独指令
    if '单独指令' in gt or '拆分' in gt:
        d = fusion_data.get('随班融合', {}).get('拆分单独指令')
        if d and d.get('pct') is not None:
            return d['pct'], '拆分单独指令'

    # 转衔跟随
    if '转衔' in gt:
        d = fusion_data.get('随班融合', {}).get('转衔跟随')
        if d and d.get('pct') is not None:
            return d['pct'], '转衔跟随'

    # 同伴发起
    if '发起' in gt and ('同伴' in gt or '社交' in gt):
        d = fusion_data.get('同伴介入', {}).get('游戏发起')
        if d and d.get('pct') is not None:
            return d['pct'], '游戏发起'

    # 同伴回应
    if '回应' in gt and '同伴' in gt:
        d = fusion_data.get('同伴介入', {}).get('游戏回应')
        if d and d.get('pct') is not None:
            return d['pct'], '游戏回应'

    # 向老师/同伴提要求
    if '提要求' in gt and ('老师' in gt or '班级' in gt):
        d = fusion_data.get('随班融合', {}).get('向老师提要求')
        if d and d.get('pct') is not None:
            return d['pct'], '向老师提要求'

    # 律动跟随（如果有）
    if '律动' in gt:
        return None, '律动跟随（反馈表未直接记录）'

    return None, None


# ═══════════════════════════════════════════════════════════════════════
# 5. Excel报告生成
# ═══════════════════════════════════════════════════════════════════════

def generate_excel(all_dates_data, dates, student_name, output_path,
                   iep_goals=None, mode='daily'):
    """生成融合数据Excel报告"""
    wb = Workbook()

    # 样式
    hf = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    hfill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cf = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
    df = Font(name='Arial', size=10)
    bf = Font(name='Arial', bold=True, size=11)
    gf = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    rf = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    yf = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    tb = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

    # ─── Sheet 1: 融合数据趋势 ───
    ws = wb.active
    ws.title = "融合数据趋势"

    ncols = 3 + len(dates) + 2  # 领域 + 指标 + 子项 + 日期列 + 均值 + 趋势
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws['A1'] = f'融合数据趋势分析 - {student_name}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
    ws['A2'] = f'分析周期：{dates[0]} 至 {dates[-1]}（共{len(dates)}天）'
    ws['A2'].font = Font(name='Arial', size=10, color='595959')

    # 表头
    row = 4
    headers = ['领域', '指标', '子项'] + [d[-5:] for d in dates] + ['均值', '趋势']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = Alignment(horizontal='center', wrap_text=True)
        c.border = tb

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 16
    for i in range(4, 4 + len(dates)):
        ws.column_dimensions[get_column_letter(i)].width = 10
    ws.column_dimensions[get_column_letter(4 + len(dates))].width = 8
    ws.column_dimensions[get_column_letter(5 + len(dates))].width = 14

    row = 5

    # ─── 个训项目部分 ───
    # 汇总所有天的个训项目名称
    all_items = OrderedDict()
    for date in dates:
        day = all_dates_data.get(date, {})
        for name, info in day.get('个训项目', {}).items():
            if name not in all_items:
                all_items[name] = {}
            all_items[name][date] = info

    if all_items:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
        ws.cell(row=row, column=1, value='【个训项目】').font = bf
        for c in range(1, ncols + 1):
            ws.cell(row=row, column=c).fill = cf; ws.cell(row=row, column=c).border = tb
        row += 1

        mastery_std = '80%'
        for date in dates:
            s = all_dates_data.get(date, {}).get('掌握标准', '80%')
            if s: mastery_std = s; break

        for name, date_info in all_items.items():
            ws.cell(row=row, column=1, value='个训').font = df; ws.cell(row=row, column=1).border = tb
            ws.cell(row=row, column=2, value=name).font = df; ws.cell(row=row, column=2).border = tb
            ws.cell(row=row, column=3, value='百分比').font = df; ws.cell(row=row, column=3).border = tb

            values = []
            mastered_any = False
            for date in dates:
                info = date_info.get(date)
                if info:
                    values.append(info.get('pct'))
                    if info.get('mastered') == '是':
                        mastered_any = True
                else:
                    values.append(None)

            for i, v in enumerate(values):
                c = ws.cell(row=row, column=4 + i)
                if v is not None:
                    c.value = v / 100; c.number_format = '0%'
                    if v >= 80: c.fill = gf
                    elif v <= 20: c.fill = rf
                else:
                    c.value = '--'; c.alignment = Alignment(horizontal='center')
                c.font = df; c.border = tb

            valid = [v for v in values if v is not None]
            avg = sum(valid) / len(valid) if valid else None
            trend = calculate_trend(values)

            ac = ws.cell(row=row, column=4 + len(dates))
            if avg is not None:
                ac.value = avg / 100; ac.number_format = '0%'
            else:
                ac.value = '--'
            ac.font = df; ac.border = tb

            trend_label = trend + (' ✓掌握' if mastered_any else '')
            tc = ws.cell(row=row, column=5 + len(dates), value=trend_label)
            tc.font = Font(name='Arial', size=10,
                           color='006100' if '上升' in trend or mastered_any else
                           '9C0006' if '下降' in trend else '000000')
            tc.border = tb

            row += 1

        # 空行分隔
        row += 1

    # 定义要追踪的融合指标
    fusion_metrics = [
        ('同伴介入', '游戏发起', 'pct'),
        ('同伴介入', '游戏回应', 'pct'),
        ('同伴介入', '影子游戏', 'pct'),
        ('同伴介入', '抓捕小鱼', 'pct'),
        ('随班融合', '集教课关注率', 'pct'),
        ('随班融合', '拆分单独指令', 'pct'),
        ('随班融合', '转衔跟随', 'pct'),
        ('随班融合', '集体指令', 'pct'),
        ('随班融合', '向老师提要求', 'pct'),
    ]

    domain_map = {
        '游戏发起': '社交', '游戏回应': '社交', '影子游戏': '社交', '抓捕小鱼': '社交',
        '集教课关注率': '集体', '拆分单独指令': '集体', '转衔跟随': '集体',
        '集体指令': '集体', '向老师提要求': '社交',
    }

    current_section = None

    for section, metric, key in fusion_metrics:
        domain = domain_map.get(metric, section)

        # Section header
        if section != current_section:
            current_section = section
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
            ws.cell(row=row, column=1, value=f'【{section}】').font = bf
            for c in range(1, ncols + 1):
                ws.cell(row=row, column=c).fill = cf; ws.cell(row=row, column=c).border = tb
            row += 1

        # Data
        ws.cell(row=row, column=1, value=domain).font = df; ws.cell(row=row, column=1).border = tb
        ws.cell(row=row, column=2, value=metric).font = df; ws.cell(row=row, column=2).border = tb
        ws.cell(row=row, column=3, value='百分比').font = df; ws.cell(row=row, column=3).border = tb

        values = []
        for date in dates:
            day = all_dates_data.get(date, {})
            val = None
            if section == '同伴介入':
                d = day.get('同伴介入', {}).get(metric)
                if d:
                    val = d.get(key)
            elif section == '随班融合':
                d = day.get('随班融合', {}).get(metric)
                if d:
                    val = d.get(key) if isinstance(d, dict) else d
            values.append(val)

        for i, v in enumerate(values):
            c = ws.cell(row=row, column=4 + i)
            if v is not None:
                c.value = v / 100; c.number_format = '0%'
                if v >= 80: c.fill = gf
                elif v <= 20: c.fill = rf
            else:
                c.value = '--'; c.alignment = Alignment(horizontal='center')
            c.font = df; c.border = tb

        valid = [v for v in values if v is not None]
        avg = sum(valid) / len(valid) if valid else None
        trend = calculate_trend(values)

        ac = ws.cell(row=row, column=4 + len(dates))
        if avg is not None:
            ac.value = avg / 100; ac.number_format = '0%'
        else:
            ac.value = '--'
        ac.font = df; ac.border = tb

        tc = ws.cell(row=row, column=5 + len(dates), value=trend)
        tc.font = Font(name='Arial', size=10,
                       color='006100' if '上升' in trend else '9C0006' if '下降' in trend else '000000')
        tc.border = tb

        row += 1

    # ─── Sheet 2: IEP达标判定（如果提供了IEP） ───
    if iep_goals:
        ws2 = wb.create_sheet("IEP达标判定")
        ws2.column_dimensions['A'].width = 10
        ws2.column_dimensions['B'].width = 50
        ws2.column_dimensions['C'].width = 12
        ws2.column_dimensions['D'].width = 12
        ws2.column_dimensions['E'].width = 12
        ws2.column_dimensions['F'].width = 15
        ws2.column_dimensions['G'].width = 15

        ws2.merge_cells('A1:G1')
        ws2['A1'] = f'IEP达标判定 - {student_name}'
        ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='1F4E79')

        iep_headers = ['领域', '目标描述', 'IEP预期', '当前均值', '数据来源', '差距', '状态']
        for col, h in enumerate(iep_headers, 1):
            c = ws2.cell(row=3, column=col, value=h)
            c.font = hf; c.fill = hfill; c.border = tb

        iep_row = 4
        for goal in iep_goals:
            # 找到这个目标在各天的匹配数据
            matched_values = []
            data_source = None
            for date in dates:
                day = all_dates_data.get(date, {})
                val, src = match_goal_to_data(goal, day)
                matched_values.append(val)
                if src:
                    data_source = src

            valid = [v for v in matched_values if v is not None]
            current_avg = sum(valid) / len(valid) if valid else None
            target = goal['target_pct']

            if current_avg is not None:
                gap = current_avg - target
                if current_avg >= target:
                    status = "已达标"
                    fill = gf
                elif current_avg >= target * 0.85:
                    status = "接近达标"
                    fill = yf
                else:
                    status = "未达标"
                    fill = rf
            else:
                gap = None
                status = "数据不足"
                fill = None

            ws2.cell(row=iep_row, column=1, value=goal['domain']).font = df
            ws2.cell(row=iep_row, column=1).border = tb

            gc = ws2.cell(row=iep_row, column=2, value=goal['goal'][:60])
            gc.font = df; gc.alignment = Alignment(wrap_text=True); gc.border = tb

            tc = ws2.cell(row=iep_row, column=3, value=f"{target}%")
            tc.font = df; tc.border = tb

            ac = ws2.cell(row=iep_row, column=4)
            if current_avg is not None:
                ac.value = f"{current_avg:.0f}%"
            else:
                ac.value = '--'
            ac.font = df; ac.border = tb

            ws2.cell(row=iep_row, column=5, value=data_source or '--').font = df
            ws2.cell(row=iep_row, column=5).border = tb

            dc = ws2.cell(row=iep_row, column=6)
            if gap is not None:
                dc.value = f"{gap:+.0f}%"
            else:
                dc.value = '--'
            dc.font = df; dc.border = tb

            sc = ws2.cell(row=iep_row, column=7, value=status)
            sc.font = df; sc.border = tb
            if fill:
                sc.fill = fill

            iep_row += 1

    # ─── Sheet 3: 文字总结 ───
    ws3 = wb.create_sheet("文字总结")
    ws3.column_dimensions['A'].width = 90
    summary = generate_text_summary(all_dates_data, dates, student_name,
                                    fusion_metrics, domain_map, iep_goals, mode)
    for i, line in enumerate(summary.split('\n'), 1):
        c = ws3.cell(row=i, column=1, value=line)
        if any(line.startswith(p) for p in ['【', '一', '二', '三', '四', '五', '六', '七']):
            c.font = Font(name='Arial', bold=True, size=11)
        else:
            c.font = Font(name='Arial', size=10)

    wb.save(output_path)
    return summary


# ═══════════════════════════════════════════════════════════════════════
# 6. 文字总结
# ═══════════════════════════════════════════════════════════════════════

def generate_text_summary(all_dates_data, dates, student_name,
                          fusion_metrics, domain_map, iep_goals=None, mode='daily'):
    lines = []
    lines.append(f"【融合数据趋势分析】")
    lines.append(f"分析周期：{dates[0]} 至 {dates[-1]}（共{len(dates)}天）")
    lines.append(f"学生：{student_name}")
    lines.append(f"模式：{'月度全局' if mode == 'monthly' else '日常趋势'}")
    lines.append("")

    # 个训项目概览
    all_items = OrderedDict()
    for date in dates:
        day = all_dates_data.get(date, {})
        for name, info in day.get('个训项目', {}).items():
            if name not in all_items:
                all_items[name] = {}
            all_items[name][date] = info

    if all_items:
        lines.append("〇、个训项目概览（融合反馈表提取）")
        for name, date_info in all_items.items():
            values = [date_info.get(d, {}).get('pct') for d in dates]
            valid = [v for v in values if v is not None]
            if not valid:
                continue
            avg = sum(valid) / len(valid)
            trend = calculate_trend(values)
            mastered = any(date_info.get(d, {}).get('mastered') == '是' for d in dates)
            vals_str = ' → '.join(f"{v}%" for v in values if v is not None)
            mark = ' ✓掌握' if mastered else ''
            lines.append(f"    {name}: {vals_str} (均值{avg:.0f}%, {trend}{mark})")
        # 统计无数据项
        no_data = [name for name, di in all_items.items()
                   if all(di.get(d, {}).get('pct') is None for d in dates)]
        if no_data:
            lines.append(f"    [未实施] {'、'.join(no_data)}")
        lines.append("")

    # 计算各指标趋势
    metric_results = []
    for section, metric, key in fusion_metrics:
        values = []
        for date in dates:
            day = all_dates_data.get(date, {})
            val = None
            if section == '同伴介入':
                d = day.get('同伴介入', {}).get(metric)
                if d: val = d.get(key)
            elif section == '随班融合':
                d = day.get('随班融合', {}).get(metric)
                if d: val = d.get(key) if isinstance(d, dict) else d
            values.append(val)

        valid = [v for v in values if v is not None]
        avg = sum(valid) / len(valid) if valid else None
        trend = calculate_trend(values)
        metric_results.append({
            'section': section, 'metric': metric,
            'domain': domain_map.get(metric, section),
            'values': values, 'avg': avg, 'trend': trend
        })

    # 一、融合数据概览
    lines.append("一、融合数据概览")
    for section in ['同伴介入', '随班融合']:
        items = [m for m in metric_results if m['section'] == section]
        if items:
            lines.append(f"  [{section}]")
            for m in items:
                valid = [v for v in m['values'] if v is not None]
                if valid:
                    vals_str = ' → '.join(f"{v}%" for v in m['values'] if v is not None)
                    lines.append(f"    {m['metric']}: {vals_str} (均值{m['avg']:.0f}%, {m['trend']})")
                else:
                    lines.append(f"    {m['metric']}: 无数据")

    # 二、进步明显
    lines.append("")
    lines.append("二、进步明显的指标（上升趋势）")
    rising = [m for m in metric_results if '上升' in m['trend']]
    if rising:
        for m in rising:
            vals_str = ' → '.join(f"{v}%" for v in m['values'] if v is not None)
            lines.append(f"  ↑ {m['metric']}（{vals_str}）")
    else:
        lines.append("  （无）")

    # 三、需要关注
    lines.append("")
    lines.append("三、需要关注的指标（下降或持续低位）")
    concern = [m for m in metric_results if '下降' in m['trend'] or '低位' in m['trend']]
    if concern:
        for m in concern:
            vals_str = ' → '.join(f"{v}%" for v in m['values'] if v is not None)
            lines.append(f"  ⚠ {m['metric']}（{vals_str}）")
    else:
        lines.append("  （无）")

    # 四、IEP达标判定
    if iep_goals:
        lines.append("")
        lines.append("四、IEP达标判定")
        for goal in iep_goals:
            matched_values = []
            for date in dates:
                day = all_dates_data.get(date, {})
                val, _ = match_goal_to_data(goal, day)
                matched_values.append(val)
            valid = [v for v in matched_values if v is not None]
            avg = sum(valid) / len(valid) if valid else None
            target = goal['target_pct']
            if avg is not None:
                gap = avg - target
                if avg >= target:
                    status = "✅ 已达标"
                elif avg >= target * 0.85:
                    status = "→ 接近达标"
                else:
                    status = "❌ 未达标"
                lines.append(f"  {status} [{goal['domain']}] {goal['goal'][:50]}（目标{target}%, 当前{avg:.0f}%, 差距{gap:+.0f}%）")
            else:
                lines.append(f"  — [{goal['domain']}] {goal['goal'][:50]}（目标{target}%, 数据不足）")

    # 五、月度全局扫描（仅月度模式）
    if mode == 'monthly':
        lines.append("")
        lines.append("五、督导观察提示清单（入园评估维度）")
        lines.append("  以下维度在每日反馈表中无法自动提取，建议在本月观察中关注：")
        manual_items = [
            "明白他人的感受", "安慰他人", "接受别人的建议",
            "接受别人的拒绝", "延迟满足", "自言自语",
            "听从区辨条件集体指令", "根据场合使用合适音量",
            "课后获得新知识", "持续独立工作", "主动打招呼",
            "恰当表达拒绝或反对", "主题活动中维持对话",
            "整理衣物", "餐后擦嘴", "整理书包", "整理桌面", "收拾玩具"
        ]
        for item in manual_items:
            lines.append(f"    □ {item}")

    return '\n'.join(lines)


# ═══════════════════════════════════════════════════════════════════════
# 7. 主流程
# ═══════════════════════════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description='ABA融合数据对比与IEP达标判定'
    )
    parser.add_argument('--files', nargs='+', required=True, help='融合反馈表PDF')
    parser.add_argument('--output', default='融合数据趋势分析.xlsx', help='输出路径')
    parser.add_argument('--student', default=None, help='学生姓名')
    parser.add_argument('--iep', default=None, help='IEP文档路径(.docx)')
    parser.add_argument('--mode', choices=['daily', 'monthly'], default='daily',
                        help='运行模式：daily=日常趋势, monthly=月度全局')
    args = parser.parse_args()

    if not check_pdftotext():
        print("错误：未找到 pdftotext。请安装：sudo apt-get install -y poppler-utils")
        sys.exit(1)

    print(f"📋 ABA融合数据对比分析")
    print(f"   输入文件：{len(args.files)} 个PDF")
    print(f"   模式：{'月度全局' if args.mode == 'monthly' else '日常趋势'}")
    if args.iep:
        print(f"   IEP文档：{args.iep}")
    print()

    # ─── 提取日期并排序 ───
    file_dates = []
    for fp in args.files:
        date = extract_date_from_filename(fp)
        if date is None:
            text = extract_layout(fp)
            date = extract_date_from_text(text)
        if date is None:
            print(f"  ⚠ 无法提取日期：{fp}，跳过")
            continue
        file_dates.append((fp, date))
    file_dates.sort(key=lambda x: x[1])

    # ─── 提取融合数据 ───
    all_dates_data = OrderedDict()
    student_name = args.student

    for fp, date in file_dates:
        print(f"  📄 处理 {os.path.basename(fp)} ({date})...")

        layout_text = extract_layout(fp)
        plain_text = extract_plain(fp)

        if student_name is None:
            student_name = extract_student_name(layout_text)

        # 个训项目（用layout模式，支持双列表格）
        items, std = parse_individual_training(layout_text)
        print(f"     个训项目：{len(items)}个（标准:{std}）")

        # 同伴介入
        peer = parse_peer_intervention(layout_text)
        peer_count = sum(1 for v in peer.values() if v.get('pct') is not None)
        print(f"     同伴介入：{peer_count}项有数据")

        # 随班融合
        inclusion = parse_inclusion_data(layout_text)
        print(f"     随班融合：{len(inclusion)}项有数据")

        # USOPAC
        usopac = parse_usopac(layout_text, student_name)

        # 每日亮点/小结
        highlight, summary = parse_daily_notes(layout_text)

        all_dates_data[date] = {
            '个训项目': items,
            '掌握标准': std,
            '同伴介入': peer,
            '随班融合': inclusion,
            'USOPAC': usopac,
            '每日亮点': highlight,
            '每日小结': summary,
        }

    if not all_dates_data:
        print("错误：没有成功提取到任何数据")
        sys.exit(1)

    if student_name is None:
        student_name = "未知学生"

    dates = list(all_dates_data.keys())

    # ─── 解析IEP（如果提供） ───
    iep_goals = None
    if args.iep:
        print(f"\n📑 解析IEP文档...")
        iep_goals = parse_iep_goals(args.iep)
        print(f"   提取到 {len(iep_goals)} 个目标")
        for g in iep_goals:
            print(f"     [{g['domain']}] {g['goal'][:50]}... (目标{g['target_pct']}%)")

    # ─── 生成报告 ───
    print(f"\n📝 生成Excel报告...")
    summary = generate_excel(
        all_dates_data, dates, student_name, args.output,
        iep_goals=iep_goals, mode=args.mode
    )

    print(f"\n✅ 报告已保存至：{args.output}")
    print(f"\n{'=' * 60}")
    print(summary)
    print(f"{'=' * 60}")


if __name__ == '__main__':
    main()
